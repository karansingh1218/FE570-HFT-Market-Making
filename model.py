#!/usr/bin/env python3
"""
SHIFT market maker: Avellaneda–Stoikov + (optional) GARCH sigma,
with a multi-tab Excel report.

FOUNDATIONAL VERSION + MINIMAL FIXES (keeps original Excel output + adds debug cols):
- Heartbeat prints (so it never looks stuck)
- stdout flush (macOS fix)
- Quote TTL + requote thresholds (mid move / sigma change)
- Non-crossing quote safety
- Top-of-book guard (avoid crossing market)
- Per-symbol gamma + min/max spread clamp (microstructure control)
- Excel workbook: same tabs + timestamped filename, with extra debug columns

Output:
  shift_run_report_<timestamp>.xlsx
"""

from __future__ import annotations

import math
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import numpy as np

# Optional: GARCH via arch. If not installed or fails, we fall back to rolling std.
try:
    from arch import arch_model  # type: ignore
    ARCH_AVAILABLE = True
except Exception:
    ARCH_AVAILABLE = False

import shift  # SHIFT Python client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


###############################################################################
# ------------------------------ CONFIG ---------------------------------------
###############################################################################

SYMBOLS = ["AAPL", "AMZN", "BRKb", "GOOG", "MSFT"]

TRADE_DURATION_MINUTES = 60

# Loop / quote management
LOOP_SLEEP_SEC = 0.35

QUOTE_TTL_SEC = 0.2      # faster refresh
REQUOTE_TICK = 0.01     # full tick move
SIGMA_REQUOTE_PCT = 0.20
SIGMA_MIN = 2e-4          # was 5e-4


# Risk limits (shares)
MAX_ABS_SHARES = 1200
REDUCE_ONLY_BAND_SHARES = 900

# Sigma model
PRICE_WINDOW = 150
GARCH_UPDATE_EVERY_N = 6
SIGMA_FALLBACK = 0.005
# SIGMA_MIN = 5e-4      # was 1e-4
SIGMA_MAX = 1.5e-2     # 1.5% per tick (safety cap)

# AS params (defaults; overridden per-symbol via SYMBOL_PARAMS)
GAMMA = 0.003
KAPPA = 1.0
HORIZON_SEC = 0.5


# Size (lots)
MIN_LOTS = 1
MAX_LOTS = 6

# Extra skew (optional)
EXTRA_SKEW_PER_LOT = 0.00

# Output
REPORT_DIR = "."

# Per-symbol microstructure clamps (you asked for these back)
# SYMBOL_PARAMS = {
#     "AAPL": dict(gamma=0.01, min_spread_c=0, max_spread_c=1),
#     "MSFT": dict(gamma=0.01, min_spread_c=0, max_spread_c=1),
#     "AMZN": dict(gamma=0.01, min_spread_c=0, max_spread_c=1),
#     "GOOG": dict(gamma=0.01, min_spread_c=0, max_spread_c=1),
#     "BRKb": dict(gamma=0.04, min_spread_c=1, max_spread_c=4),
# }

SYMBOL_PARAMS = {
    "AAPL": dict(gamma=0.005, min_spread_c=0, max_spread_c=0),
    "MSFT": dict(gamma=0.005, min_spread_c=0, max_spread_c=0),
    "AMZN": dict(gamma=0.005, min_spread_c=0, max_spread_c=0),
    "GOOG": dict(gamma=0.005, min_spread_c=0, max_spread_c=0),
    "BRKb": dict(gamma=0.01,  min_spread_c=0, max_spread_c=1),
}




###############################################################################
# ------------------------------ UTIL -----------------------------------------
###############################################################################

def now_utc() -> datetime:
    return datetime.utcnow()

def ts_str(ts: datetime) -> str:
    return ts.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

def clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))

def round_tick(px: float, tick: float = 0.01) -> float:
    return round(round(px / tick) * tick, 2)

def safe_mid_from_best(best: Any, last_mid: Optional[float]) -> Tuple[float, float, float]:
    """Compute mid with fallbacks. Returns (bid, ask, mid)."""
    bid = float(getattr(best, "get_bid_price")())
    ask = float(getattr(best, "get_ask_price")())
    if bid > 0 and ask > 0:
        return bid, ask, (bid + ask) / 2.0
    if bid > 0:
        return bid, ask, bid
    if ask > 0:
        return bid, ask, ask
    if last_mid is not None and last_mid > 0:
        return bid, ask, last_mid
    return bid, ask, 0.01


###############################################################################
# ------------------------------ SIGMA ----------------------------------------
###############################################################################

def clamp(x, lo, hi):
    return max(lo, min(hi, x))

def estimate_sigma_garch(prices, window=300):
    try:
        px = np.asarray(prices, dtype=float)
        px = px[np.isfinite(px)]
        px = px[px > 0]

        if px.size < max(30, window + 1):
            if px.size < 2:
                return SIGMA_MIN
            rets = np.diff(np.log(px))
            rets = rets[np.isfinite(rets)]
            s = np.nanstd(rets) if rets.size else SIGMA_MIN
            return clamp(s, SIGMA_MIN, SIGMA_MAX)

        # recent log returns
        log_px = np.log(px)
        rets = np.diff(log_px)[-window:]
        rets = rets[np.isfinite(rets)]

        if rets.size < 30:
            s = np.nanstd(rets) if rets.size else SIGMA_MIN
            return clamp(s, SIGMA_MIN, SIGMA_MAX)

        # scale returns for stability
        scale = 100.0
        r = rets * scale

        am = arch_model(
            r,
            mean="Zero",
            vol="GARCH",
            p=1, q=1,
            dist="normal"
        )
        res = am.fit(disp="off")

        opt = getattr(res, "optimization_result", None)
        if opt is not None and hasattr(opt, "success"):
            if not opt.success:
                s = np.nanstd(rets)
                return clamp(s, SIGMA_MIN, SIGMA_MAX)

        f = res.forecast(horizon=1, reindex=False)
        var1 = float(f.variance.values[-1, 0])

        if not np.isfinite(var1) or var1 <= 0:
            s = np.nanstd(rets)
            return clamp(s, SIGMA_MIN, SIGMA_MAX)

        sigma = np.sqrt(var1) / scale
        return clamp(float(sigma), SIGMA_MIN, SIGMA_MAX)

    except Exception:
        return SIGMA_MIN

def estimate_sigma_garch_old(log_returns: np.ndarray) -> float:
    """GARCH(1,1) on log returns. Falls back to rolling std."""
    rets = np.asarray(log_returns, dtype=float)
    rets = rets[np.isfinite(rets)]
    rets = rets[np.abs(rets) > 1e-12]
    if len(rets) < 20:
        s = float(np.std(rets)) if len(rets) > 1 else SIGMA_FALLBACK
        return clamp(s, SIGMA_MIN, SIGMA_MAX)

    if not ARCH_AVAILABLE:
        s = float(np.std(rets))
        return clamp(s, SIGMA_MIN, SIGMA_MAX)

    scaled = rets * 100.0
    try:
        am = arch_model(
            scaled,
            vol="Garch",
            p=1,
            q=1,
            mean="Zero",
            dist="normal",
            rescale=False,
        )
        res = am.fit(disp="off")
        f = res.forecast(horizon=1)
        var_scaled = float(f.variance.iloc[-1, 0])
        sigma_scaled = math.sqrt(max(var_scaled, 0.0))
        sigma = sigma_scaled / 100.0
        return clamp(float(sigma), SIGMA_MIN, SIGMA_MAX)
    except Exception:
        s = float(np.std(rets))
        return clamp(s, SIGMA_MIN, SIGMA_MAX)


###############################################################################
# ------------------------------ AS QUOTES ------------------------------------
###############################################################################

def compute_as_quotes(
    mid: float,
    inv_lots: float,
    sigma: float,
    gamma: float,
    kappa: float,
    horizon_sec: float,
    min_spread_c: int,
    max_spread_c: int,
) -> Tuple[float, float, float, float, float, float, float, float]:
    """
    Return:
      bid, ask, reservation_price, raw_half, clamped_half, min_half, max_half, spread
    """
    q = float(inv_lots)
    T = float(horizon_sec)

    r = mid - q * gamma * (sigma ** 2) * T
    r -= q * EXTRA_SKEW_PER_LOT

    spread = gamma * (sigma ** 2) * T + 2.0 * math.log(1.0 + gamma / kappa)
    raw_half = spread / 2.0

    min_half = (min_spread_c / 100.0) / 2.0
    max_half = (max_spread_c / 100.0) / 2.0
    half = clamp(raw_half, min_half, max_half)

    bid = round_tick(max(0.01, r - half))
    ask = round_tick(max(bid + 0.01, r + half))

    return bid, ask, r, raw_half, half, min_half, max_half, spread


###############################################################################
# ------------------------------ SIZING ---------------------------------------
###############################################################################

def size_from_inventory(inv_shares: int) -> Tuple[int, int]:
    """Return (bid_lots, ask_lots), inventory-aware and risk-limited."""
    inv_lots = inv_shares / 100.0

    bid_mult = math.exp(-0.15 * max(inv_lots, 0.0))
    ask_mult = math.exp(-0.15 * max(-inv_lots, 0.0))

    bid_lots = int(round(MAX_LOTS * bid_mult))
    ask_lots = int(round(MAX_LOTS * ask_mult))

    bid_lots = int(clamp(bid_lots, MIN_LOTS, MAX_LOTS))
    ask_lots = int(clamp(ask_lots, MIN_LOTS, MAX_LOTS))

    # Reduce-only band
    if inv_shares > REDUCE_ONLY_BAND_SHARES:
        bid_lots = 0
        ask_lots = MAX_LOTS
    elif inv_shares < -REDUCE_ONLY_BAND_SHARES:
        bid_lots = MAX_LOTS
        ask_lots = 0

    # Hard risk stop
    if inv_shares >= MAX_ABS_SHARES:
        bid_lots = 0
    if inv_shares <= -MAX_ABS_SHARES:
        ask_lots = 0

    return bid_lots, ask_lots

def lots_to_shares(lots: int) -> int:
    return int(lots) * 100

def ceil_lots_from_shares(shares: int) -> int:
    return int(math.ceil(shares / 100.0))


###############################################################################
# ------------------------------ LOGGER ---------------------------------------
###############################################################################

@dataclass
class RunLogs:
    market: List[Dict[str, Any]] = field(default_factory=list)
    quotes: List[Dict[str, Any]] = field(default_factory=list)
    orders: List[Dict[str, Any]] = field(default_factory=list)
    fills: List[Dict[str, Any]] = field(default_factory=list)
    inventory: List[Dict[str, Any]] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class SymbolState:
    sym: str
    mids: List[float] = field(default_factory=list)
    sigma: float = SIGMA_FALLBACK
    ewma: Any = None
    seeded: bool = False

    last_mid: Optional[float] = None
    last_bid: Optional[float] = None
    last_ask: Optional[float] = None
    loops: int = 0

    last_quote_ts: Optional[float] = None
    last_quote_bid: Optional[float] = None
    last_quote_ask: Optional[float] = None
    last_quote_bid_lots: int = 0
    last_quote_ask_lots: int = 0

    max_long_shares: int = 0
    max_short_shares: int = 0
    max_abs_shares: int = 0

    initial_realized_pl: float = 0.0
    processed_exec_keys: set = field(default_factory=set)


###############################################################################
# ------------------------------ ORDER MGMT -----------------------------------
###############################################################################

class MarketMaker:
    def __init__(self):
        self.total_fills = 0

    def on_execution(self, execution):
        self.total_fills += execution.get_size()


def cancel_resting_symbol_orders(trader: shift.Trader, sym: str, logs: RunLogs) -> int:
    """Cancel all waiting orders for this symbol only."""
    count = 0
    try:
        for o in list(trader.get_waiting_list()):
            if getattr(o, "symbol", None) == sym:
                trader.submit_cancellation(o)
                logs.orders.append({
                    "ts": ts_str(now_utc()),
                    "symbol": sym,
                    "action": "CANCEL",
                    "order_id": getattr(o, "id", None),
                    "type": str(getattr(o, "type", "")),
                    "lots": getattr(o, "size", None),
                    "price": getattr(o, "price", None),
                })
                count += 1
                time.sleep(0.03)
    except Exception as e:
        logs.errors.append({"ts": ts_str(now_utc()), "symbol": sym, "where": "cancel_resting", "error": repr(e)})
    return count


def close_positions_robust(trader: shift.Trader, sym: str, logs: RunLogs, max_tries: int = 12) -> None:
    """Flatten positions using CEIL(lots) and retry until flat."""
    for _ in range(max_tries):
        try:
            item = trader.get_portfolio_item(sym)
            long_sh = int(item.get_long_shares())
            short_sh = int(item.get_short_shares())

            if long_sh == 0 and short_sh == 0:
                return

            if long_sh > 0:
                lots = ceil_lots_from_shares(long_sh)
                order = shift.Order(shift.Order.Type.MARKET_SELL, sym, lots)
                trader.submit_order(order)
                logs.orders.append({
                    "ts": ts_str(now_utc()), "symbol": sym, "action": "FLATTEN_SELL",
                    "order_id": getattr(order, "id", None), "type": "MARKET_SELL",
                    "lots": lots, "price": None,
                })

            if short_sh > 0:
                lots = ceil_lots_from_shares(short_sh)
                order = shift.Order(shift.Order.Type.MARKET_BUY, sym, lots)
                trader.submit_order(order)
                logs.orders.append({
                    "ts": ts_str(now_utc()), "symbol": sym, "action": "FLATTEN_BUY",
                    "order_id": getattr(order, "id", None), "type": "MARKET_BUY",
                    "lots": lots, "price": None,
                })

            time.sleep(0.6)
        except Exception as e:
            logs.errors.append({"ts": ts_str(now_utc()), "symbol": sym, "where": "close_positions", "error": repr(e)})
            time.sleep(0.6)

    logs.errors.append({"ts": ts_str(now_utc()), "symbol": sym, "where": "close_positions", "error": "Not flat after retries"})


###############################################################################
# ------------------------------ FILL LOGGING ---------------------------------
###############################################################################

def process_new_fills(trader: shift.Trader, sym: str, st: SymbolState, logs: RunLogs) -> None:
    """Scan submitted orders and fetch new executions for this symbol."""
    try:
        submitted = list(trader.get_submitted_orders())
    except Exception as e:
        logs.errors.append({"ts": ts_str(now_utc()), "symbol": sym, "where": "get_submitted_orders", "error": repr(e)})
        return

    for o in submitted:
        try:
            if getattr(o, "symbol", None) != sym:
                continue
            oid = getattr(o, "id", None)
            if oid is None:
                continue

            executed_size = getattr(o, "executed_size", 0)
            if not executed_size or executed_size <= 0:
                continue

            try:
                execs = list(trader.get_executed_orders(oid))
            except Exception:
                execs = []

            for ex in execs:
                ex_size = getattr(ex, "executed_size", 0)
                ex_px = getattr(ex, "executed_price", None)
                ex_type = str(getattr(ex, "type", ""))
                if not ex_size or ex_size <= 0 or ex_px is None:
                    continue

                ex_time = getattr(ex, "timestamp", None)
                ex_time_str = str(ex_time) if ex_time is not None else ts_str(now_utc())

                key = (oid, ex_time_str, float(ex_px), int(ex_size), ex_type)
                if key in st.processed_exec_keys:
                    continue
                st.processed_exec_keys.add(key)

                side = "BUY" if "BUY" in ex_type else "SELL"
                lots = int(ex_size)
                shares = lots_to_shares(lots)

                logs.fills.append({
                    "ts": ex_time_str,
                    "symbol": sym,
                    "side": side,
                    "lots": lots,
                    "shares": shares,
                    "price": float(ex_px),
                    "order_id": oid,
                    "order_type": str(getattr(o, "type", "")),
                    "mid_at_fill": st.last_mid,
                    "quoted_bid": st.last_quote_bid,
                    "quoted_ask": st.last_quote_ask,
                    "sigma": st.sigma,
                })
        except Exception as e:
            logs.errors.append({"ts": ts_str(now_utc()), "symbol": sym, "where": "process_new_fills", "error": repr(e)})


###############################################################################
# ------------------------------ MTM ESTIMATE ---------------------------------
###############################################################################

def compute_unrealized_mtm(item: Any, bid: float, ask: float) -> float:
    """Estimate unrealized P&L from best bid/ask and avg position prices."""
    try:
        long_sh = int(item.get_long_shares())
        short_sh = int(item.get_short_shares())
        long_px = float(item.get_long_price())
        short_px = float(item.get_short_price())
    except Exception:
        return 0.0

    if bid <= 0 and ask > 0:
        bid = ask
    if ask <= 0 and bid > 0:
        ask = bid
    if bid <= 0 or ask <= 0:
        return 0.0

    unreal_long = long_sh * (bid - long_px) if long_sh > 0 else 0.0
    unreal_short = short_sh * (short_px - ask) if short_sh > 0 else 0.0
    return float(unreal_long + unreal_short)


###############################################################################
# ------------------------------ EXCEL REPORT ---------------------------------
###############################################################################

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            s = str(v)
            widths[i] = max(widths.get(i, 8), min(60, len(s) + 2))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, headers: List[str], rows: List[List[Any]], freeze: str = "A2") -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for r in rows:
        ws.append(r)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(headers)), max(1, len(rows) + 1))
    autosize_columns(ws)


def build_excel_report(
    outfile: str,
    symbols: List[str],
    start_ts: datetime,
    end_ts: datetime,
    params: Dict[str, Any],
    per_sym: Dict[str, SymbolState],
    logs: RunLogs,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("RunInfo")
    info_rows = [
        ["start_utc", ts_str(start_ts)],
        ["end_utc", ts_str(end_ts)],
        ["duration_min", (end_ts - start_ts).total_seconds() / 60.0],
        ["symbols", ", ".join(symbols)],
    ]
    for k, v in params.items():
        info_rows.append([k, v])
    write_table(ws, ["key", "value"], info_rows, freeze="A2")

    # Market (keep original cols + add debug cols at end)
    ws = wb.create_sheet("Market")
    m_headers = [
        "ts", "symbol",
        "best_bid", "best_ask", "mid", "mkt_spread",
        "sigma",
        "shares", "long_shares", "short_shares",
        "avg_long_px", "avg_short_px",
        "realized_pl", "unreal_mtm", "total_pl_est",
        # debug extras
        "loop", "ttl_ok", "moved", "sigma_changed", "should_quote",
    ]
    m_rows = [[r.get(h) for h in m_headers] for r in logs.market]
    write_table(ws, m_headers, m_rows)

    # Quotes (keep original cols + add debug cols at end)
    ws = wb.create_sheet("Quotes")
    q_headers = [
        "ts", "symbol",
        "mid", "sigma", "inv_shares", "inv_lots",
        "bid_px", "bid_lots", "ask_px", "ask_lots",
        "quoted_spread", "reason", "canceled_count",
        "reservation_px", "half_spread",
        # debug extras
        "gamma", "kappa", "T", "raw_half", "clamped_half", "min_half", "max_half",
        "mid_move", "sigma_move", "ttl_ok",
    ]
    q_rows = [[r.get(h) for h in q_headers] for r in logs.quotes]
    write_table(ws, q_headers, q_rows)

    ws = wb.create_sheet("Orders")
    o_headers = ["ts", "symbol", "action", "order_id", "type", "lots", "price"]
    o_rows = [[r.get(h) for h in o_headers] for r in logs.orders]
    write_table(ws, o_headers, o_rows)

    ws = wb.create_sheet("Fills")
    f_headers = [
        "ts", "symbol", "side",
        "lots", "shares", "price",
        "order_id", "order_type",
        "mid_at_fill", "slippage_to_mid",
        "cashflow", "sigma", "quoted_bid", "quoted_ask",
    ]
    f_rows = []
    for r in logs.fills:
        mid = r.get("mid_at_fill")
        px = r.get("price")
        slip = (float(px) - float(mid)) if (mid not in (None, 0) and px is not None) else None
        side = r.get("side")
        shares = r.get("shares")
        cash = None
        if side and shares and px is not None:
            cash = (-1 if side == "BUY" else 1) * int(shares) * float(px)
        f_rows.append([
            r.get("ts"), r.get("symbol"), side,
            r.get("lots"), shares, px,
            r.get("order_id"), r.get("order_type"),
            mid, slip,
            cash, r.get("sigma"), r.get("quoted_bid"), r.get("quoted_ask"),
        ])
    write_table(ws, f_headers, f_rows)

    ws = wb.create_sheet("Inventory")
    i_headers = [
        "ts", "symbol",
        "shares", "long_shares", "short_shares",
        "avg_long_px", "avg_short_px",
        "realized_pl", "unreal_mtm", "total_pl_est",
        "max_long", "max_short", "max_abs",
    ]
    i_rows = [[r.get(h) for h in i_headers] for r in logs.inventory]
    write_table(ws, i_headers, i_rows)

    ws = wb.create_sheet("Summary")
    s_headers = [
        "symbol",
        "initial_realized_pl",
        "final_realized_pl",
        "net_realized_pl",
        "fills_count",
        "fill_shares",
        "buy_shares",
        "sell_shares",
        "avg_buy_px",
        "avg_sell_px",
        "max_abs_shares",
    ]
    rows = []
    for sym in symbols:
        st = per_sym[sym]
        rows.append([sym, st.initial_realized_pl, None, None, None, None, None, None, None, None, st.max_abs_shares])
    write_table(ws, s_headers, rows)

    for r in range(2, 2 + len(symbols)):
        sym_cell = "A{}".format(r)
        ws["E{}".format(r)] = "=COUNTIF(Fills!$B:$B,{})".format(sym_cell)
        ws["F{}".format(r)] = "=SUMIF(Fills!$B:$B,{},Fills!$E:$E)".format(sym_cell)
        ws["G{}".format(r)] = "=SUMIFS(Fills!$E:$E,Fills!$B:$B,{},Fills!$C:$C,\"BUY\")".format(sym_cell)
        ws["H{}".format(r)] = "=SUMIFS(Fills!$E:$E,Fills!$B:$B,{},Fills!$C:$C,\"SELL\")".format(sym_cell)
        ws["I{}".format(r)] = (
            "=IFERROR(SUMPRODUCT((Fills!$B:$B={})*(Fills!$C:$C=\"BUY\")*Fills!$E:$E*Fills!$F:$F)"
            "/SUMIFS(Fills!$E:$E,Fills!$B:$B,{},Fills!$C:$C,\"BUY\"),\"\")"
        ).format(sym_cell, sym_cell)
        ws["J{}".format(r)] = (
            "=IFERROR(SUMPRODUCT((Fills!$B:$B={})*(Fills!$C:$C=\"SELL\")*Fills!$E:$E*Fills!$F:$F)"
            "/SUMIFS(Fills!$E:$E,Fills!$B:$B,{},Fills!$C:$C,\"SELL\"),\"\")"
        ).format(sym_cell, sym_cell)
        ws["D{}".format(r)] = "=C{}-B{}".format(r, r)

    last_realized: Dict[str, float] = {s: per_sym[s].initial_realized_pl for s in symbols}
    seen = set()
    for row in reversed(logs.inventory):
        sym = row.get("symbol")
        if sym in symbols and sym not in seen:
            val = row.get("realized_pl")
            if val is not None:
                last_realized[sym] = float(val)
            seen.add(sym)
            if len(seen) == len(symbols):
                break
    for idx, sym in enumerate(symbols, start=2):
        ws["C{}".format(idx)] = last_realized.get(sym)

    ws = wb.create_sheet("Errors")
    e_headers = ["ts", "symbol", "where", "error"]
    e_rows = [[r.get(h) for h in e_headers] for r in logs.errors]
    write_table(ws, e_headers, e_rows)

    wb.save(outfile)


class EWMASigma:
    def __init__(self, halflife_sec=10.0):
        self.halflife = float(halflife_sec)
        self.last_price = None
        self.last_ts = None
        self.var = None

    def update(self, price: float, now_ts: float) -> float:
        if price <= 0 or not np.isfinite(price):
            return self.current()

        if self.last_price is None:
            self.last_price = price
            self.last_ts = now_ts
            return self.current()

        dt = max(1e-6, now_ts - self.last_ts)
        lam = math.exp(-math.log(2.0) * dt / self.halflife)

        r = math.log(price / self.last_price)
        if np.isfinite(r):
            if self.var is None:
                self.var = r * r
            else:
                self.var = lam * self.var + (1.0 - lam) * (r * r)
            # print(
            #     f"[VOL_DBG] r={r:.6e} var={self.var:.6e}",
            #     flush=True
            # )

        self.last_price = price
        self.last_ts = now_ts
        return self.current()

    def current(self) -> float:
        if self.var is None:
            return SIGMA_FALLBACK
        return clamp(math.sqrt(max(self.var, 0.0)), SIGMA_MIN, SIGMA_MAX)


###############################################################################
# ------------------------------ MAIN STRATEGY --------------------------------
###############################################################################
def run(trader: shift.Trader) -> str:
    logs = RunLogs()
    states: Dict[str, SymbolState] = {s: SymbolState(sym=s) for s in SYMBOLS}

    # EWMA volatility
    for st in states.values():
        st.ewma = EWMASigma(halflife_sec=60.0)

    # initial realized P&L
    for sym in SYMBOLS:
        try:
            states[sym].initial_realized_pl = float(
                trader.get_portfolio_item(sym).get_realized_pl()
            )
        except Exception:
            states[sym].initial_realized_pl = 0.0

    start_utc = now_utc()

    # market-time stop
    try:
        last_trade_time = trader.get_last_trade_time()
        end_time = last_trade_time + timedelta(minutes=TRADE_DURATION_MINUTES)
        print("[RUN] last_trade_time =", last_trade_time, flush=True)
        print("[RUN] end_time       =", end_time, flush=True)
    except Exception:
        end_time = now_utc() + timedelta(minutes=TRADE_DURATION_MINUTES)

    # wall-clock safety (10 min max)
    wall_start = time.time()
    WALL_CLOCK_MAX_SEC = 3 * 60 * 60  # 3 hours

    trader.sub_all_order_book()
    time.sleep(1.0)

    last_hb = 0.0

    while True:
        # ---------------- HEARTBEAT ----------------
        now_wall = time.time()
        if now_wall - last_hb >= 1.0:
            print("[HEARTBEAT]", ts_str(now_utc()), flush=True)
            last_hb = now_wall

        # ---------------- STOP CONDITIONS ----------------
        if time.time() - wall_start >= WALL_CLOCK_MAX_SEC:
            print("[STOP] wall clock", flush=True)
            break

        try:
            if trader.get_last_trade_time() >= end_time:
                print("[STOP] market time", flush=True)
                break
        except Exception:
            pass

        loop_start = time.time()

        for sym in SYMBOLS:
            st = states[sym]
            st.loops += 1

            # ---------- Best prices ----------
            try:
                best = trader.get_best_price(sym)
                bid, ask, mid = safe_mid_from_best(best, st.last_mid)
                st.last_mid = mid
            except Exception:
                continue

            # ---------- Portfolio ----------
            try:
                item = trader.get_portfolio_item(sym)
                shares = int(item.get_shares())
                realized = float(item.get_realized_pl())
            except Exception:
                continue
            # ---------- TERMINAL LIQUIDATION (FINITE HORIZON) ----------
            try:
                time_left_sec = (end_time - trader.get_last_trade_time()).total_seconds()
            except Exception:
                time_left_sec = 9999  # fallback

            TERMINAL_WINDOW_SEC = 60  # last 1 minute
            terminal_phase = time_left_sec <= TERMINAL_WINDOW_SEC

            if terminal_phase and shares != 0 and bid > 0 and ask > 0:
                print(f"[{sym}] TERMINAL LIQUIDATION inv={shares}", flush=True)

                # Cancel resting quotes
                cancel_resting_symbol_orders(trader, sym, logs)

                if shares > 0:
                    # Sell inventory
                    lots = min(ceil_lots_from_shares(shares), MAX_LOTS)
                    trader.submit_order(
                        shift.Order(shift.Order.Type.MARKET_SELL, sym, lots)
                    )
                else:
                    # Buy to cover
                    lots = min(ceil_lots_from_shares(-shares), MAX_LOTS)
                    trader.submit_order(
                        shift.Order(shift.Order.Type.MARKET_BUY, sym, lots)
                    )

                # Skip normal quoting this loop
                continue

            inv_lots = shares / 100.0
            st.max_abs_shares = max(st.max_abs_shares, abs(shares))

            # ---------- Sigma ----------
            sigma_before = st.sigma
            st.sigma = st.ewma.update(mid, time.time())

            sigma = st.sigma

            # ---------- Quote decision ----------
            now_s = time.time()
            ttl_ok = st.last_quote_ts is None or (now_s - st.last_quote_ts) >= QUOTE_TTL_SEC

            moved = True
            if st.last_quote_bid and st.last_quote_ask:
                last_mid = (st.last_quote_bid + st.last_quote_ask) / 2.0
                moved = abs(mid - last_mid) >= REQUOTE_TICK

            sigma_changed = (
                sigma_before > 0
                and abs(sigma - sigma_before) / sigma_before >= SIGMA_REQUOTE_PCT
            )

            should_quote = ttl_ok and (moved or sigma_changed or st.last_quote_ts is None)

            # ---------- Market log ----------
            logs.market.append({
                "ts": ts_str(now_utc()),
                "symbol": sym,
                "best_bid": bid,
                "best_ask": ask,
                "mid": mid,
                "mkt_spread": ask - bid if bid > 0 and ask > 0 else None,
                "sigma": sigma,
                "shares": shares,
                "realized_pl": realized,
                "loop": st.loops,
                "ttl_ok": ttl_ok,
                "moved": moved,
                "sigma_changed": sigma_changed,
                "should_quote": should_quote,
            })

            process_new_fills(trader, sym, st, logs)

            if not should_quote or mid <= 0:
                continue

            canceled = cancel_resting_symbol_orders(trader, sym, logs)

            # ---------- ONE-TIME SEED ----------
            if not st.seeded and shares == 0 and bid > 0 and ask > 0:
                trader.submit_order(
                    shift.Order(shift.Order.Type.MARKET_BUY, sym, 1)
                )
                st.seeded = True
                st.last_quote_ts = now_s
                print(f"[{sym}] SEED BUY", flush=True)
                continue

            # ---------- AS pricing ----------
            sp = SYMBOL_PARAMS.get(sym, {})
            gamma = float(sp.get("gamma", GAMMA))
            min_c = int(sp.get("min_spread_c", 0))
            max_c = int(sp.get("max_spread_c", 0))

            bid_px, ask_px, r_px, raw_half, half, min_half, max_half, _ = compute_as_quotes(
                mid, inv_lots, sigma, gamma, KAPPA, HORIZON_SEC, min_c, max_c
            )

            # ---------- FORCE TOP-OF-BOOK ----------
            if bid > 0 and ask > 0:
                bid_px = max(bid_px, bid)
                ask_px = min(ask_px, ask)

            # ---------- FORCED AGGRESSION ----------

            # ==========================================================
            # FORCED AGGRESSION (HORIZON LIQUIDATION)
            # ==========================================================
            FORCE_EVERY_N_LOOPS = 20  # ~20 seconds (0.35s loop)

            forced = False
            bid_lots, ask_lots = size_from_inventory(shares)

            if st.loops % FORCE_EVERY_N_LOOPS == 0:
                forced = True

                if shares > 0:
                    # Force SELL to unwind inventory
                    ask_px = round_tick(bid)
                    bid_lots = 0
                    ask_lots = min(2, MAX_LOTS)
                    print(f"[{sym}] FORCE SELL @ {ask_px}", flush=True)

                elif shares < 0:
                    # Force BUY to cover
                    bid_px = round_tick(ask)
                    bid_lots = min(2, MAX_LOTS)
                    ask_lots = 0
                    print(f"[{sym}] FORCE BUY @ {bid_px}", flush=True)

                else:
                    # Flat → seed small buy
                    bid_px = round_tick(ask)
                    bid_lots = 1
                    ask_lots = 0
                    print(f"[{sym}] FORCE SEED BUY @ {bid_px}", flush=True)

            # FORCE_EVERY_N_LOOPS = 300  # ~20 seconds
            # forced = False
            # if st.loops % FORCE_EVERY_N_LOOPS == 0:
            #     bid_px = round_tick(ask)
            #     ask_px = bid_px + 0.01
            #     bid_lots = min(2, MAX_LOTS)
            #     ask_lots = 0
            #     forced = True
            # else:
            #     bid_lots, ask_lots = size_from_inventory(shares)

            # ---------- RECORD QUOTES (THIS IS WHAT YOU WERE MISSING) ----------
            logs.quotes.append({
                "ts": ts_str(now_utc()),
                "symbol": sym,
                "mid": mid,
                "sigma": sigma,
                "inv_shares": shares,
                "inv_lots": inv_lots,
                "bid_px": bid_px,
                "bid_lots": bid_lots,
                "ask_px": ask_px,
                "ask_lots": ask_lots,
                "quoted_spread": ask_px - bid_px,
                "reason": "FORCED" if forced else "NORMAL",
                "canceled_count": canceled,
                "reservation_px": r_px,
                "half_spread": half,
                "gamma": gamma,
                "kappa": KAPPA,
                "T": HORIZON_SEC,
                "raw_half": raw_half,
                "clamped_half": half,
                "min_half": min_half,
                "max_half": max_half,
                "mid_move": abs(mid - (st.last_mid or mid)),
                "sigma_move": abs(sigma - sigma_before),
                "ttl_ok": ttl_ok,
            })

            print(
                f"[{sym}] quote=({bid_px:.2f},{ask_px:.2f}) "
                f"lots=({bid_lots},{ask_lots}) inv={shares} sigma={sigma:.5f}",
                flush=True,
            )

            # ---------- SUBMIT ----------
            if bid_lots > 0:
                trader.submit_order(
                    shift.Order(shift.Order.Type.LIMIT_BUY, sym, bid_lots, bid_px)
                )
            if ask_lots > 0:
                trader.submit_order(
                    shift.Order(shift.Order.Type.LIMIT_SELL, sym, ask_lots, ask_px)
                )

            st.last_quote_ts = now_s
            st.last_quote_bid = bid_px
            st.last_quote_ask = ask_px

        time.sleep(max(0.0, LOOP_SLEEP_SEC - (time.time() - loop_start)))

    # ---------- CLEANUP ----------
    for sym in SYMBOLS:
        cancel_resting_symbol_orders(trader, sym, logs)
        close_positions_robust(trader, sym, logs)

    end_utc = now_utc()
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    outfile = f"shift_run_report_{stamp}.xlsx"

    build_excel_report(outfile, SYMBOLS, start_utc, end_utc, {}, states, logs)
    return outfile

# def run(trader: shift.Trader) -> str:
#     logs = RunLogs()
#     states: Dict[str, SymbolState] = {s: SymbolState(sym=s) for s in SYMBOLS}
#
#     # EWMA volatility (stable, fast)
#     for s in states.values():
#         s.ewma = EWMASigma(halflife_sec=60.0)
#
#     # Record initial realized P&L
#     for sym in SYMBOLS:
#         try:
#             item = trader.get_portfolio_item(sym)
#             states[sym].initial_realized_pl = float(item.get_realized_pl())
#         except Exception:
#             states[sym].initial_realized_pl = 0.0
#
#     start_time_utc = now_utc()
#
#     # Market-time horizon
#     try:
#         last_trade_time = trader.get_last_trade_time()
#         end_time = last_trade_time + timedelta(minutes=TRADE_DURATION_MINUTES)
#         print(f"[RUN] last_trade_time={last_trade_time}", flush=True)
#         print(f"[RUN] end_time       ={end_time}", flush=True)
#     except Exception:
#         end_time = now_utc() + timedelta(minutes=TRADE_DURATION_MINUTES)
#
#     WALL_CLOCK_MAX_SEC = 10 * 60   # 10 minutes max runtime
#     wall_start = time.time()
#
#     try:
#         trader.sub_all_order_book()
#         time.sleep(1.0)
#     except Exception as e:
#         logs.errors.append({
#             "ts": ts_str(now_utc()),
#             "symbol": "*",
#             "where": "sub_all_order_book",
#             "error": repr(e),
#         })
#
#     last_hb = 0.0
#
#     while True:
#         # ---------------- HEARTBEAT ----------------
#         now_wall = time.time()
#         if now_wall - last_hb >= 1.0:
#             print("[HEARTBEAT]", ts_str(now_utc()), flush=True)
#             last_hb = now_wall
#
#         # ---------------- HARD STOP ----------------
#         if time.time() - wall_start >= WALL_CLOCK_MAX_SEC:
#             print("[STOP] Wall-clock limit reached", flush=True)
#             break
#
#         # ---------------- MARKET-TIME STOP ----------
#         try:
#             if trader.get_last_trade_time() >= end_time:
#                 print("[STOP] Market-time limit reached", flush=True)
#                 break
#         except Exception:
#             pass
#
#         loop_start = time.time()
#
#         for sym in SYMBOLS:
#             st = states[sym]
#             st.loops += 1
#
#             # ---------- Best prices ----------
#             try:
#                 best = trader.get_best_price(sym)
#                 bid, ask, mid = safe_mid_from_best(best, st.last_mid)
#                 st.last_mid, st.last_bid, st.last_ask = mid, bid, ask
#             except Exception:
#                 continue
#
#             # ---------- Portfolio ----------
#             try:
#                 item = trader.get_portfolio_item(sym)
#                 shares = int(item.get_shares())
#                 long_sh = int(item.get_long_shares())
#                 short_sh = int(item.get_short_shares())
#                 avg_long = float(item.get_long_price())
#                 avg_short = float(item.get_short_price())
#                 realized = float(item.get_realized_pl())
#             except Exception:
#                 continue
#
#             inv_lots = shares / 100.0
#             st.max_abs_shares = max(st.max_abs_shares, abs(shares))
#
#             unreal = compute_unrealized_mtm(item, bid, ask)
#             total_est = realized + unreal
#
#             # ---------- Sigma (EWMA) ----------
#             sigma_before = st.sigma
#             st.sigma = st.ewma.update(mid, time.time())
#             sigma = st.sigma
#
#             # ---------- Quote decision ----------
#             now_s = time.time()
#             ttl_ok = st.last_quote_ts is None or (now_s - st.last_quote_ts) >= QUOTE_TTL_SEC
#
#             moved = True
#             if st.last_quote_bid and st.last_quote_ask:
#                 last_mid = (st.last_quote_bid + st.last_quote_ask) / 2.0
#                 moved = abs(mid - last_mid) >= REQUOTE_TICK
#
#             sigma_changed = (
#                 sigma_before > 0
#                 and abs(sigma - sigma_before) / sigma_before >= SIGMA_REQUOTE_PCT
#             )
#
#             should_quote = ttl_ok and (moved or sigma_changed or st.last_quote_ts is None)
#
#             # ---------- Log market ----------
#             logs.market.append({
#                 "ts": ts_str(now_utc()),
#                 "symbol": sym,
#                 "best_bid": bid,
#                 "best_ask": ask,
#                 "mid": mid,
#                 "mkt_spread": ask - bid if bid > 0 and ask > 0 else None,
#                 "sigma": sigma,
#                 "shares": shares,
#                 "long_shares": long_sh,
#                 "short_shares": short_sh,
#                 "avg_long_px": avg_long,
#                 "avg_short_px": avg_short,
#                 "realized_pl": realized,
#                 "unreal_mtm": unreal,
#                 "total_pl_est": total_est,
#                 "loop": st.loops,
#                 "ttl_ok": ttl_ok,
#                 "moved": moved,
#                 "sigma_changed": sigma_changed,
#                 "should_quote": should_quote,
#             })
#
#             process_new_fills(trader, sym, st, logs)
#
#             if not should_quote or mid <= 0:
#                 continue
#
#             canceled = cancel_resting_symbol_orders(trader, sym, logs)
#
#             # ---------- PRE-SEED INVENTORY ----------
#             # ---------- PRE-SEED INVENTORY (ONCE ONLY) ----------
#             if not st.seeded and shares == 0 and bid > 0 and ask > 0:
#                 trader.submit_order(
#                     shift.Order(shift.Order.Type.MARKET_BUY, sym, 1)
#                 )
#
#                 st.seeded = True
#                 st.last_quote_ts = time.time()
#
#                 print(f"[{sym}] SEED BUY submitted", flush=True)
#                 continue
#
#             sp = SYMBOL_PARAMS.get(sym, {})
#             gamma = float(sp.get("gamma", GAMMA))
#             min_c = int(sp.get("min_spread_c", 1))
#             max_c = int(sp.get("max_spread_c", 10))
#
#             bid_px, ask_px, r_px, raw_half, half, min_half, max_half, _ = compute_as_quotes(
#                 mid, inv_lots, sigma, gamma, KAPPA, HORIZON_SEC, min_c, max_c
#             )
#
#             # ==========================================================
#             # (5) FORCE TOP-OF-BOOK (ABSOLUTELY REQUIRED IN SHIFT)
#             # ==========================================================
#             if bid > 0 and ask > 0:
#                 bid_px = max(bid_px, bid)  # do not sit behind best bid
#                 ask_px = min(ask_px, ask)  # do not sit behind best ask
#
#             # ==========================================================
#             # (6) FORCED AGGRESSION (GUARANTEES FILLS)
#             # ==========================================================
#             FORCE_EVERY_N_LOOPS = 60  # ~20 seconds
#
#             if st.loops % FORCE_EVERY_N_LOOPS == 0:
#                 # Cross the spread to GUARANTEE at least one fill
#                 bid_px = round_tick(ask)
#                 ask_px = round_tick(bid_px + 0.01)
#
#                 bid_lots = min(2, MAX_LOTS)
#                 ask_lots = 0
#
#                 print(f"[{sym}] FORCE BUY @ {bid_px}", flush=True)
#             else:
#                 bid_lots, ask_lots = size_from_inventory(shares)
#
#             # bid_lots, ask_lots = size_from_inventory(shares)
#
#             print(
#                 f"[{sym}] quote=({bid_px:.2f},{ask_px:.2f}) "
#                 f"lots=({bid_lots},{ask_lots}) inv={shares} sigma={sigma:.5f}",
#                 flush=True,
#             )
#
#             if bid_lots > 0:
#                 trader.submit_order(
#                     shift.Order(shift.Order.Type.LIMIT_BUY, sym, bid_lots, bid_px)
#                 )
#
#             if ask_lots > 0:
#                 trader.submit_order(
#                     shift.Order(shift.Order.Type.LIMIT_SELL, sym, ask_lots, ask_px)
#                 )
#
#             st.last_quote_ts = now_s
#             st.last_quote_bid = bid_px
#             st.last_quote_ask = ask_px
#
#         time.sleep(max(0.0, LOOP_SLEEP_SEC - (time.time() - loop_start)))
#
#     # ---------- CLEANUP ----------
#     for sym in SYMBOLS:
#         cancel_resting_symbol_orders(trader, sym, logs)
#         close_positions_robust(trader, sym, logs)
#
#     end_time_utc = now_utc()
#     stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
#     outfile = os.path.join(REPORT_DIR, f"shift_run_report_{stamp}.xlsx")
#
#     build_excel_report(outfile, SYMBOLS, start_time_utc, end_time_utc, {}, states, logs)
#     return outfile

def main() -> None:
    username = os.getenv("SHIFT_USERNAME", "ksingh29")
    # DO NOT hardcode passwords. Set SHIFT_PASSWORD in your shell:
    # export SHIFT_PASSWORD="..."
    password = os.getenv("SHIFT_PASSWORD", "")
    cfg = os.getenv("SHIFT_CFG", "initiator.cfg")

    print("[SHIFT] user={} cfg={}".format(username, cfg), flush=True)

    with shift.Trader(username) as trader:
        trader.connect(cfg, password)
        print("[INFO] Connection established.", flush=True)
        time.sleep(1.0)

        report = run(trader)
        print("[DONE] Report written: {}".format(report), flush=True)


if __name__ == "__main__":
    main()
